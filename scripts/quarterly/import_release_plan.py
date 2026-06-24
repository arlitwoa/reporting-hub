#!/usr/bin/env python3
"""Import sprint calendar from EPC Release Planning.xlsx; engine releases from PDE Jira."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from artifact.atlassian import AtlassianAdapter  # noqa: E402

from extensions.twoa_programme.delivery_health import load_delivery_health_config  # noqa: E402
from extensions.twoa_programme.pde_engine_releases import build_quarter_engine_calendar  # noqa: E402
from extensions.twoa_programme.quarterly_reporting import load_quarterly_reporting_config  # noqa: E402
from extensions.twoa_programme.release_plan_calendar import extend_sprint_calendar  # noqa: E402
from scripts.quarterly.common import CONFIG_PATH, out_path  # noqa: E402

DEFAULT_XLSX = Path.home() / "Downloads" / "EPC Release Planning.xlsx"
DEFAULT_SEED = _REPO_ROOT / "config" / "quarterly" / "seeds" / "release-plan-metadata.json"
SHEET = "Release Plan"
SPRINT_ROW = 3
DATE_ROW = 2
_SPRINT_STATIC_KEYS = frozenset(
    {"name", "sprintNumber", "startDate", "endDate", "projected"}
)


def _parse_sprints_from_workbook(
    path: Path,
    *,
    quarter_start: date,
    quarter_end: date,
) -> tuple[list[dict], str | None]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET!r} not found; have: {wb.sheetnames}")
    ws = wb[SHEET]
    row2 = list(ws.iter_rows(min_row=DATE_ROW, max_row=DATE_ROW, values_only=True))[0]
    row3 = list(ws.iter_rows(min_row=SPRINT_ROW, max_row=SPRINT_ROW, values_only=True))[0]

    starts = [i for i, value in enumerate(row3) if value and "Sprint" in str(value)]
    sprints: list[dict] = []
    for index, start_col in enumerate(starts):
        end_col = starts[index + 1] - 1 if index + 1 < len(starts) else start_col + 13
        start_dt = row2[start_col] if start_col < len(row2) else None
        end_dt = row2[min(end_col, len(row2) - 1)] if row2 else None
        if not isinstance(start_dt, datetime) or not isinstance(end_dt, datetime):
            continue
        sprint_start = start_dt.date()
        sprint_end = end_dt.date()
        if sprint_end < quarter_start or sprint_start > quarter_end:
            continue
        sprints.append(
            {
                "name": str(row3[start_col]).strip(),
                "startDate": sprint_start.isoformat(),
                "endDate": sprint_end.isoformat(),
            }
        )

    calendar_start = next(
        (row2[i].date().isoformat() for i, v in enumerate(row2) if isinstance(v, datetime)),
        None,
    )
    return sprints, calendar_start


def _fetch_pde_releases(
    quarter_start: date,
    quarter_end: date,
) -> dict:
    health = load_delivery_health_config()
    risk = health.dev_done_risk
    if risk is None:
        raise SystemExit("delivery-health.json devDoneRisk section required for PDE engine releases.")
    adapter = AtlassianAdapter.from_profile(
        "atlassian", profiles_dir=os.environ["ARTIFACT_PROFILES_DIR"]
    )
    return build_quarter_engine_calendar(
        adapter,
        quarter_start=quarter_start,
        quarter_end=quarter_end,
        engine_project_key=risk.engine_project_key,
        engine_issue_type=risk.engine_issue_type,
        engine_in_cycle_filter=risk.engine_in_cycle_filter,
    )


def _strip_sprint_row(row: dict) -> dict:
    return {key: row[key] for key in _SPRINT_STATIC_KEYS if key in row}


def build_release_plan_from_seed(
    *,
    seed: Path,
    quarter_start: date,
    quarter_end: date,
    releases_from_jira: bool = True,
) -> dict:
    """Use committed sprint calendar; refresh engine releases from Jira when available."""
    payload = json.loads(seed.read_text(encoding="utf-8"))
    sprints = [_strip_sprint_row(row) for row in payload.get("sprints") or []]
    sprints = extend_sprint_calendar(sprints, quarter_start=quarter_start, quarter_end=quarter_end)

    release_payload: dict | None = None
    if releases_from_jira and os.environ.get("ARTIFACT_PROFILES_DIR"):
        release_payload = _fetch_pde_releases(quarter_start, quarter_end)
        releases = release_payload["inCycleReleases"]
    else:
        releases = payload.get("inCycleReleases") or []

    return {
        "sourceFile": str(seed),
        "sheet": payload.get("sheet", SHEET),
        "sprintRow": payload.get("sprintRow", SPRINT_ROW),
        "dateRow": payload.get("dateRow", DATE_ROW),
        "calendarStartsAt": payload.get("calendarStartsAt"),
        "quarterStart": quarter_start.isoformat(),
        "quarterEnd": quarter_end.isoformat(),
        "sprints": sprints,
        "inCycleReleases": releases,
        "releaseSource": release_payload.get("source") if release_payload else payload.get("releaseSource"),
        "inCycleFixVersion": (
            release_payload.get("inCycleFixVersion") if release_payload else payload.get("inCycleFixVersion")
        ),
        "releaseCadenceDays": (
            release_payload.get("releaseCadenceDays") if release_payload else payload.get("releaseCadenceDays")
        ),
        "extended": True,
    }


def build_release_plan_metadata(
    *,
    xlsx: Path,
    quarter_start: date,
    quarter_end: date,
    releases_from_jira: bool = True,
) -> dict:
    sprints, calendar_start = _parse_sprints_from_workbook(
        xlsx,
        quarter_start=quarter_start,
        quarter_end=quarter_end,
    )
    sprints = extend_sprint_calendar(sprints, quarter_start=quarter_start, quarter_end=quarter_end)

    release_payload: dict | None = None
    if releases_from_jira and os.environ.get("ARTIFACT_PROFILES_DIR"):
        release_payload = _fetch_pde_releases(quarter_start, quarter_end)
        releases = release_payload["inCycleReleases"]
    else:
        releases = []

    return {
        "sourceFile": str(xlsx),
        "sheet": SHEET,
        "sprintRow": SPRINT_ROW,
        "dateRow": DATE_ROW,
        "calendarStartsAt": calendar_start,
        "quarterStart": quarter_start.isoformat(),
        "quarterEnd": quarter_end.isoformat(),
        "sprints": sprints,
        "inCycleReleases": releases,
        "releaseSource": release_payload.get("source") if release_payload else None,
        "inCycleFixVersion": release_payload.get("inCycleFixVersion") if release_payload else None,
        "releaseCadenceDays": release_payload.get("releaseCadenceDays") if release_payload else None,
        "extended": True,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import sprint calendar from Release Plan xlsx and PDE engine releases from Jira."
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Path to EPC Release Planning.xlsx")
    parser.add_argument(
        "--seed",
        type=Path,
        default=DEFAULT_SEED,
        help="Sprint calendar seed when the workbook is unavailable (CI / GitHub Actions)",
    )
    parser.add_argument("--sprints-only", action="store_true", help="Skip PDE Jira release fetch")
    parser.add_argument("--write", action="store_true", help="Write release-plan-metadata.json")
    args = parser.parse_args(argv)

    config = load_quarterly_reporting_config(CONFIG_PATH)
    if args.xlsx.is_file():
        payload = build_release_plan_metadata(
            xlsx=args.xlsx,
            quarter_start=config.quarter.start_date,
            quarter_end=config.quarter.end_date,
            releases_from_jira=not args.sprints_only,
        )
    elif args.seed.is_file():
        payload = build_release_plan_from_seed(
            seed=args.seed,
            quarter_start=config.quarter.start_date,
            quarter_end=config.quarter.end_date,
            releases_from_jira=not args.sprints_only,
        )
    else:
        raise SystemExit(f"Workbook not found: {args.xlsx} and no seed at {args.seed}")
    text = json.dumps(payload, indent=2)
    print(text)

    if args.write:
        path = out_path("release-plan-metadata.json")
        path.write_text(text + "\n", encoding="utf-8")
        print(f"Wrote {path}", file=sys.stderr)
        if payload.get("inCycleReleases"):
            engine_path = out_path("engine-releases.json")
            engine_rows = [
                {
                    "name": row["name"],
                    "releaseDate": row["releaseDate"],
                    **({"carriageType": row["carriageType"]} if row.get("carriageType") else {}),
                    **({"projected": True} if row.get("projected") else {}),
                }
                for row in payload["inCycleReleases"]
            ]
            engine_path.write_text(json.dumps(engine_rows, indent=2) + "\n", encoding="utf-8")
            print(f"Wrote {engine_path}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
